#!/usr/bin/env python3
"""
process_data.py
Reads portwatch CSV from data/ folder, outputs:
  - data/dashboard_data.json  (used by index.html)
  - data/hormuz_analysis.xlsx (downloadable Excel with 4 sheets + 2 charts)
"""
import pandas as pd
import numpy as np
import json
import os
import sys
from pathlib import Path
from datetime import datetime

ROOT     = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"

# ── Find CSV ───────────────────────────────────────────────────
csv_files = list(DATA_DIR.glob("*.csv"))
if not csv_files:
    print("ERROR: No CSV file found in data/"); sys.exit(1)
CSV_PATH = sorted(csv_files)[-1]
print(f"Processing: {CSV_PATH.name}")

# ── Load ───────────────────────────────────────────────────────
df = pd.read_csv(CSV_PATH, encoding='utf-8-sig')
df['date'] = pd.to_datetime(df['date'], utc=True, errors='coerce').dt.tz_localize(None)
df = df.sort_values(['portname','date']).reset_index(drop=True)

CHOKEPOINTS = ['Strait of Hormuz', 'Cape of Good Hope', 'Suez Canal']
WAR_DATE    = pd.Timestamp('2026-02-28')
START_DATE  = pd.Timestamp('2026-01-01')
LATEST_DATE = df['date'].max()
dates       = pd.date_range(START_DATE, LATEST_DATE, freq='D')
date_strs   = [d.strftime('%Y-%m-%d') for d in dates]

# ── Daily n_total ──────────────────────────────────────────────
daily = {}
for cp in CHOKEPOINTS:
    sub = df[df['portname']==cp].set_index('date')['n_total'].sort_index().astype(float)
    daily[cp] = sub.reindex(dates)

# ── 7-day centred MA ──────────────────────────────────────────
ma7 = {}
for cp in CHOKEPOINTS:
    full = df[df['portname']==cp].set_index('date')['n_total'].sort_index().astype(float)
    ma_full = full.rolling(7, center=True, min_periods=1).mean()
    ma7[cp] = ma_full.reindex(dates)

# ── Stats ──────────────────────────────────────────────────────
stats = {}
for cp in CHOKEPOINTS:
    pre  = [float(daily[cp].get(d)) for d in dates
            if d < WAR_DATE and pd.notna(daily[cp].get(d))]
    post = [float(daily[cp].get(d)) for d in dates
            if d >= WAR_DATE and pd.notna(daily[cp].get(d))]
    peak_val  = float(daily[cp].max())
    peak_date = daily[cp].idxmax()
    stats[cp] = {
        'pre_avg':    round(np.mean(pre),  1) if pre  else 0,
        'post_avg':   round(np.mean(post), 1) if post else 0,
        'latest':     int(daily[cp].iloc[-1]) if pd.notna(daily[cp].iloc[-1]) else 0,
        'peak':       int(peak_val),
        'peak_date':  peak_date.strftime('%d %b') if pd.notna(peak_date) else '',
        'pct_change': round((np.mean(post)/np.mean(pre)-1)*100, 1)
                      if pre and np.mean(pre) > 0 else 0,
    }

# ── Write JSON ─────────────────────────────────────────────────
data_out = {
    'meta': {
        'updated':      LATEST_DATE.strftime('%Y-%m-%d'),
        'updated_fmt':  LATEST_DATE.strftime('%B %d, %Y'),
        'source':       'IMF PortWatch · AIS Satellite Data',
        'war_date':     '2026-02-28',
        'war_date_fmt': 'February 28, 2026',
        'n_days':       len(dates),
        'generated':    datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'),
    },
    'dates': date_strs,
    'daily': {
        cp: [int(daily[cp].get(d)) if pd.notna(daily[cp].get(d)) else 0
             for d in dates]
        for cp in CHOKEPOINTS
    },
    'ma7': {
        cp: [round(float(ma7[cp].get(d)), 2) if pd.notna(ma7[cp].get(d)) else 0
             for d in dates]
        for cp in CHOKEPOINTS
    },
    'stats': stats,
}

json_path = DATA_DIR / 'dashboard_data.json'
with open(json_path, 'w') as f:
    json.dump(data_out, f, indent=2)
print(f"✅ JSON → {json_path}")

# ── Write Excel ────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    DARK   = PatternFill("solid", fgColor="1A1F2E")
    HORMUZ = PatternFill("solid", fgColor="C0392B")
    CAPE   = PatternFill("solid", fgColor="1A5276")
    SUEZ   = PatternFill("solid", fgColor="B7770D")
    LIGHT  = PatternFill("solid", fgColor="F0F4FA")
    ALT    = PatternFill("solid", fgColor="FAFBFD")
    WAR_F  = PatternFill("solid", fgColor="FDECEA")
    WAR_H  = PatternFill("solid", fgColor="FDCECE")
    t      = Side(style='thin', color='D0D7E3')
    THIN   = Border(left=t, right=t, top=t, bottom=t)
    FWH    = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    FG     = Font(name='Calibri', italic=True, color='5A657A', size=9)
    CTR    = Alignment(horizontal='center', vertical='center')
    WRP    = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def st(c, font=None, fill=None, align=None, border=None, fmt=None):
        if font:   c.font = font
        if fill:   c.fill = fill
        if align:  c.alignment = align
        if border: c.border = border
        if fmt:    c.number_format = fmt

    wb = Workbook()
    N  = len(dates)
    DS = 3   # data starts row 3 in both sheets

    def make_data_sheet(ws, title, col_fmt, use_formula=False, ds_ref_ws=None):
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        st(ws['A1'], font=FWH, fill=DARK, align=CTR)
        ws.row_dimensions[1].height = 26
        for ci2, (h, f) in enumerate(zip(
            ['Date','Strait of Hormuz','Cape of Good Hope','Suez Canal','War Period'],
            [DARK, HORMUZ, CAPE, SUEZ, DARK]), 1):
            cell = ws.cell(row=2, column=ci2, value=h)
            st(cell, font=FWH, fill=f, align=WRP, border=THIN)
        ws.row_dimensions[2].height = 28
        cp_clrs = ['C0392B','1A5276','B7770D']
        for ri2, date in enumerate(dates, DS):
            is_war   = date >= WAR_DATE
            is_first = date == WAR_DATE
            bg = WAR_H if is_first else (WAR_F if is_war else (LIGHT if ri2%2==0 else ALT))
            cell_a = ws.cell(row=ri2, column=1, value=date.to_pydatetime().date())
            st(cell_a,
               font=Font(name='Calibri', bold=is_first, size=10,
                         color='C0392B' if is_first else ('882222' if is_war else '1A1F2E')),
               fill=bg, align=CTR, border=THIN, fmt='DD-MMM-YYYY')
            for ci2b, (col_l, cp, clr) in enumerate(
                    zip(['B','C','D'], CHOKEPOINTS, cp_clrs), 2):
                cell = ws.cell(row=ri2, column=ci2b)
                if use_formula:
                    row_lo = max(DS + (ri2-DS) - 3, DS)
                    row_hi = min(DS + (ri2-DS) + 3, DS + N - 1)
                    cell.value = (f'=IFERROR(AVERAGE({ds_ref_ws}!{col_l}{row_lo}'
                                  f':{ds_ref_ws}!{col_l}{row_hi}),"")')
                    fmt2 = '0.00'
                else:
                    val = int(daily[cp].get(date)) if pd.notna(daily[cp].get(date)) else 0
                    cell.value = val
                    fmt2 = '0'
                st(cell,
                   font=Font(name='Calibri', bold=is_war, size=10,
                             color=clr if is_war else '333333'),
                   fill=bg, align=CTR, border=THIN, fmt=fmt2)
            cell_e = ws.cell(row=ri2, column=5, value='WAR' if is_war else '')
            st(cell_e,
               font=Font(name='Calibri', bold=is_first, size=9,
                         color='C0392B' if is_war else 'AAAAAA'),
               fill=bg, align=CTR, border=THIN)
            ws.row_dimensions[ri2].height = 16
        for c2, w in [('A',14),('B',20),('C',20),('D',16),('E',10)]:
            ws.column_dimensions[c2].width = w
        ws.freeze_panes = 'A3'
        return DS + N - 1

    # Sheet 1 & 2 — data
    ws1 = wb.active; ws1.title = "Daily_Ships"
    DS_END = make_data_sheet(ws1,
        f'Daily Ships — Jan 2026 to {LATEST_DATE.strftime("%b %d, %Y")}  |  IMF PortWatch',
        '0', use_formula=False)

    ws2 = wb.create_sheet("MA7_Ships")
    MS = DS   # MA7_Ships data also starts at row DS=3
    MA_END = make_data_sheet(ws2,
        f'7-Day MA — Jan 2026 to {LATEST_DATE.strftime("%b %d, %Y")}  |  IMF PortWatch',
        '0.00', use_formula=True, ds_ref_ws='Daily_Ships')

    # Sheets 3 & 4 — charts
    def make_chart_sheet(ws, title, sub, src_ws_name, src_start, src_end,
                         y_max, y_unit, y_fmt, smooth=True):
        ws.merge_cells('A1:M1')
        ws['A1'] = title
        st(ws['A1'], font=Font(name='Calibri', bold=True, color='1A1F2E', size=13),
           fill=PatternFill("solid", fgColor="EBF0FA"), align=CTR)
        ws.row_dimensions[1].height = 28
        ws.merge_cells('A2:M2')
        ws['A2'] = sub
        st(ws['A2'], font=FG, fill=PatternFill("solid", fgColor="F5F8FD"), align=CTR)
        ws.row_dimensions[2].height = 16
        ws.merge_cells('A3:M3')
        ws['A3'] = '🚨  Feb 28, 2026 — US-Israel War on Iran: IRGC declares Hormuz CLOSED'
        st(ws['A3'], font=Font(name='Calibri', bold=True, color='C0392B', size=10),
           fill=PatternFill("solid", fgColor="FDECEA"),
           align=Alignment(horizontal='left', vertical='center'))
        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 8

        for ci2, (h, f) in enumerate(zip(
            ['Date','Strait of Hormuz','Cape of Good Hope','Suez Canal','War Start'],
            [DARK, HORMUZ, CAPE, SUEZ, DARK]), 1):
            cell = ws.cell(row=5, column=ci2, value=h)
            st(cell, font=FWH, fill=f, align=WRP, border=THIN)
        ws.row_dimensions[5].height = 28

        CS = 6
        for ri2, date in enumerate(dates, CS):
            is_war   = date >= WAR_DATE
            is_first = date == WAR_DATE
            bg = WAR_H if is_first else (WAR_F if is_war else (LIGHT if ri2%2==0 else ALT))
            cell_a = ws.cell(row=ri2, column=1, value=date.strftime('%d-%b'))
            st(cell_a, font=Font(name='Calibri', bold=is_first, size=9,
                                 color='C0392B' if is_first else '444444'),
               fill=bg, align=CTR, border=THIN)
            src_row = src_start + (ri2 - CS)
            for ci2b, col_l in enumerate(['B','C','D'], 2):
                cell = ws.cell(row=ri2, column=ci2b)
                cell.value = f'={src_ws_name}!{col_l}{src_row}'
                st(cell, fill=bg, align=CTR, border=THIN,
                   fmt='0' if y_fmt == '0' else '0.00')
            cell_e = ws.cell(row=ri2, column=5)
            if is_first:
                cell_e.value = y_max
                st(cell_e, font=Font(name='Calibri', bold=True, color='E74C3C', size=9),
                   fill=WAR_H, align=CTR, border=THIN)
            else:
                cell_e.value = None
                st(cell_e, fill=bg, border=THIN)
            ws.row_dimensions[ri2].height = 14
        CE = CS + N - 1

        ch = LineChart()
        ch.title = None
        ch.y_axis.title = 'Ships / Day' + ('' if smooth else ' (7-day MA)')
        ch.x_axis.title = f'Date (Jan – {LATEST_DATE.strftime("%b %Y")})'
        ch.style = 10; ch.width = 24; ch.height = 14
        ch.y_axis.scaling.min = 0; ch.y_axis.scaling.max = y_max
        ch.y_axis.numFmt = y_fmt; ch.y_axis.majorUnit = y_unit
        ch.legend.position = 'b'
        ch.set_categories(Reference(ws, min_col=1, min_row=CS, max_row=CE))
        for ci2b, (clr, lw) in enumerate(
                [('C0392B',28000),('1A5276',24000),('B7770D',24000)], 2):
            ref = Reference(ws, min_col=ci2b, min_row=CS-1, max_row=CE)
            ch.add_data(ref, titles_from_data=True)
            s = ch.series[ci2b-2]
            s.graphicalProperties.line.solidFill = clr
            s.graphicalProperties.line.width = lw
            s.smooth = smooth; s.marker.symbol = 'none'
        ref_w = Reference(ws, min_col=5, min_row=CS-1, max_row=CE)
        ch.add_data(ref_w, titles_from_data=True)
        ch.series[3].graphicalProperties.line.solidFill = 'E74C3C'
        ch.series[3].graphicalProperties.line.width = 20000
        ch.series[3].graphicalProperties.line.dashDot = 'dash'
        ch.series[3].smooth = False; ch.series[3].marker.symbol = 'none'
        ch.series[3].title = SeriesLabel(v='War start (Feb 28)')
        ws.add_chart(ch, 'G1')
        for col in range(1,6): ws.column_dimensions[get_column_letter(col)].width = 11
        for col in range(7,22): ws.column_dimensions[get_column_letter(col)].width = 9

    ws3 = wb.create_sheet("Chart_Daily")
    make_chart_sheet(ws3,
        'Daily Ships — Hormuz vs Cape of Good Hope vs Suez Canal',
        f'Jan 2026 – {LATEST_DATE.strftime("%b %d, %Y")}  ·  n_total/day  ·  IMF PortWatch AIS',
        'Daily_Ships', DS, DS_END, 160, 20, '0', smooth=True)

    ws4 = wb.create_sheet("Chart_MA7")
    make_chart_sheet(ws4,
        '7-Day Moving Average — Hormuz vs Cape of Good Hope vs Suez Canal',
        f'Jan 2026 – {LATEST_DATE.strftime("%b %d, %Y")}  ·  7-day centred MA  ·  IMF PortWatch AIS',
        'MA7_Ships', MS, MA_END, 140, 20, '0.0', smooth=True)

    wb._sheets = [ws1, ws2, ws3, ws4]
    xlsx_path = DATA_DIR / 'hormuz_analysis.xlsx'
    wb.save(xlsx_path)
    print(f"✅ Excel → {xlsx_path}  ({os.path.getsize(xlsx_path)//1024} KB)")

except Exception as e:
    import traceback; traceback.print_exc()
    print(f"⚠ Excel failed: {e}")

print(f"\n✅ Done — {len(dates)} days up to {LATEST_DATE.strftime('%B %d, %Y')}")
